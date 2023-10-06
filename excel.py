import openpyxl

# Specify the file path of the Excel file
loc = "C:\\Users\\User\\Desktop\\Python Programs\\practice2.xlsx"

# Load the Excel file
wb = openpyxl.load_workbook(loc)

# Select the first sheet in the Excel file
sh = wb.worksheets[0]

# Print the value of cell (3, 3) in the sheet
print(sh.cell(row=3, column=3).value)

print(sh.max_row)
print(sh.max_column)

#to extract elements from the sheet
for i in range(sh.max_column):  # length of the column
    cell_value = sh.cell(row=1, column=i+1).value
    print(cell_value)


# Extract column values
column_index = 1  # Specify the column index you want to extract (e.g., 1 for column A)
for i in range(1,sh.max_row + 1):
    cell_value = sh.cell(row=i, column=1).value
    print(cell_value)


#print row values
row_values = list(sh.cell(row=1,column=5).value)
print(row_values)